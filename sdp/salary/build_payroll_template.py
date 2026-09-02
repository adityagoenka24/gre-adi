from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
INPUT_FONT = Font(name=FONT_NAME, color="0000FF", size=11)
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=11)
LINK_FONT = Font(name=FONT_NAME, color="008000", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F3864")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="808080")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
N_EMP = 10  # number of employee slots

wb = Workbook()

def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------------- Sheet 1: Employee Master ----------------
ws = wb.active
ws.title = "Employee Master"
ws["A1"] = "S.D. Pharmaceuticals — Employee Master"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:D1")
ws["A2"] = "Fill in one row per employee. Basic = fixed monthly amount paid in full. Daily Allowance = variable amount paid per day actually present."
ws["A2"].font = NOTE_FONT
ws.merge_cells("A2:D2")

headers = ["Employee Name", "Basic (Monthly, ₹)", "Daily Allowance (₹/day)", "Notes"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

for r in range(5, 5 + N_EMP):
    for c in range(1, 5):
        ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=r, column=2).font = INPUT_FONT
    ws.cell(row=r, column=2).number_format = '#,##0;(#,##0);"-"'
    ws.cell(row=r, column=3).font = INPUT_FONT
    ws.cell(row=r, column=3).number_format = '#,##0;(#,##0);"-"'
    ws.cell(row=r, column=1).font = INPUT_FONT
    ws.cell(row=r, column=4).font = INPUT_FONT

set_widths(ws, [22, 18, 20, 26])
ws.freeze_panes = "A5"

# ---------------- Sheet 2: Monthly Working Days ----------------
ws2 = wb.create_sheet("Monthly Working Days")
ws2["A1"] = "Total Working Days per Month"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:B1")
ws2["A2"] = 'Enter Month as "Mon-YYYY" (e.g. Jul-2026). Add a new row each month.'
ws2["A2"].font = NOTE_FONT
ws2.merge_cells("A2:C2")

for i, h in enumerate(["Month", "Total Working Days", "Notes"], start=1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, 3)

ws2["A5"] = "Jul-2026"
ws2["B5"] = 26
for r in range(5, 25):
    for c in range(1, 4):
        ws2.cell(row=r, column=c).border = BORDER
    ws2.cell(row=r, column=1).font = INPUT_FONT
    ws2.cell(row=r, column=2).font = INPUT_FONT
    ws2.cell(row=r, column=3).font = INPUT_FONT

set_widths(ws2, [14, 20, 30])
ws2.freeze_panes = "A5"

# ---------------- Sheet 3: Attendance Log (absences only) ----------------
ws3 = wb.create_sheet("Attendance Log")
ws3["A1"] = "Attendance Log — log ONLY the days an employee was ABSENT"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:D1")
ws3["A2"] = "Every working day, add one row per absent employee. Present days are calculated automatically (Working Days − Absences)."
ws3["A2"].font = NOTE_FONT
ws3.merge_cells("A2:D2")

for i, h in enumerate(["Date", "Employee Name", "Month (auto)", "Reason (optional)"], start=1):
    ws3.cell(row=4, column=i, value=h)
style_header(ws3, 4, 4)

MAX_LOG_ROWS = 400
for r in range(5, 5 + MAX_LOG_ROWS):
    ws3.cell(row=r, column=1).number_format = "DD-MMM-YYYY"
    ws3.cell(row=r, column=1).font = INPUT_FONT
    ws3.cell(row=r, column=2).font = INPUT_FONT
    ws3.cell(row=r, column=3).value = f'=IF(A{r}="","",TEXT(A{r},"MMM-YYYY"))'
    ws3.cell(row=r, column=3).font = FORMULA_FONT
    ws3.cell(row=r, column=4).font = INPUT_FONT
    for c in range(1, 5):
        ws3.cell(row=r, column=c).border = BORDER

dv_emp = DataValidation(type="list", formula1=f"='Employee Master'!$A$5:$A${4+N_EMP}", allow_blank=True)
ws3.add_data_validation(dv_emp)
dv_emp.add(f"B5:B{4+MAX_LOG_ROWS}")

set_widths(ws3, [14, 22, 14, 26])
ws3.freeze_panes = "A5"

# ---------------- Sheet 4: FY Opening Leaves ----------------
ws5 = wb.create_sheet("FY Opening Leaves")
ws5["A1"] = "FY Opening Leave Balances"
ws5["A1"].font = TITLE_FONT
ws5.merge_cells("A1:C1")
ws5["A2"] = 'Leaves already taken BEFORE this tracker started, per employee per FY (Apr-Mar). FY Label = "FY" + year the FY starts (e.g. Apr 2026-Mar 2027 = "FY2026"). Leave at 0 if not applicable; add a new FY block each year if needed.'
ws5["A2"].font = NOTE_FONT
ws5.merge_cells("A2:C2")

for i, h in enumerate(["FY Label", "Employee Name", "Opening Leaves"], start=1):
    ws5.cell(row=4, column=i, value=h)
style_header(ws5, 4, 3)

FY_OPEN_ROWS = 60
for r in range(5, 5 + FY_OPEN_ROWS):
    for c in range(1, 4):
        ws5.cell(row=r, column=c).border = BORDER
    ws5.cell(row=r, column=1).font = INPUT_FONT
    ws5.cell(row=r, column=2).font = INPUT_FONT
    ws5.cell(row=r, column=3).font = INPUT_FONT

dv_emp2 = DataValidation(type="list", formula1=f"='Employee Master'!$A$5:$A${4+N_EMP}", allow_blank=True)
ws5.add_data_validation(dv_emp2)
dv_emp2.add(f"B5:B{4+FY_OPEN_ROWS}")

EMPLOYEES_SEED = ["Samir Das", "Sujit Das", "Bishnu Das", "Soumen Roy", "Mangal Bannerjee", "Partha Das"]
for i, name in enumerate(EMPLOYEES_SEED):
    ws5.cell(row=5 + i, column=1, value="FY2026")
    ws5.cell(row=5 + i, column=2, value=name)
    ws5.cell(row=5 + i, column=3, value=0)

set_widths(ws5, [12, 22, 16])
ws5.freeze_panes = "A5"

# ---------------- Sheet 5: Monthly Payslip Summary ----------------
ws4 = wb.create_sheet("Monthly Payslip Summary")
ws4["A1"] = "Monthly Payslip Summary"
ws4["A1"].font = TITLE_FONT
ws4.merge_cells("A1:M1")
ws4["A3"] = "Month:"
ws4["A3"].font = Font(name=FONT_NAME, bold=True)
ws4["B3"] = "Jul-2026"
ws4["B3"].font = INPUT_FONT
ws4["B3"].fill = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
ws4["C3"] = "← change this each month, then read the rows below"
ws4["C3"].font = NOTE_FONT

# Helper calc block (month/FY date boundaries), off to the side
ws4["O2"] = "Helper dates (do not edit)"
ws4["O2"].font = NOTE_FONT
helpers = [
    ("O3", "Month Start", "P3", '=DATEVALUE("01-"&$B$3)'),
    ("O4", "Month End", "P4", "=EOMONTH($P$3,0)"),
    ("O5", "FY Start", "P5", "=IF(MONTH($P$3)>=4,DATE(YEAR($P$3),4,1),DATE(YEAR($P$3)-1,4,1))"),
    ("O6", "Prev Month End", "P6", "=$P$3-1"),
    ("O7", "FY Label", "P7", '="FY"&YEAR($P$5)'),
]
for lbl_cell, lbl, val_cell, formula in helpers:
    ws4[lbl_cell] = lbl
    ws4[lbl_cell].font = Font(name=FONT_NAME, size=9, color="808080")
    ws4[val_cell] = formula
    ws4[val_cell].font = FORMULA_FONT
for c in ["P3", "P4", "P5", "P6"]:
    ws4[c].number_format = "DD-MMM-YYYY"

headers4 = ["Employee Name", "Basic (₹)", "Daily Allowance (₹/day)",
            "Total Working Days", "Days Absent", "Days Present", "Variable Payout (₹)",
            "FY Leaves Used\nBefore This Month", "FY Leaves Used\nIncl. This Month",
            "Excess Leaves\n(beyond 30/FY)", "Basic Deduction (₹)", "Adjusted Basic (₹)", "Total Payout (₹)"]
for i, h in enumerate(headers4, start=1):
    ws4.cell(row=5, column=i, value=h)
style_header(ws4, 5, len(headers4))

last_master_row = 4 + N_EMP
for idx, r in enumerate(range(6, 6 + N_EMP)):
    mrow = 5 + idx
    ws4.cell(row=r, column=1, value=f"='Employee Master'!A{mrow}").font = LINK_FONT
    ws4.cell(row=r, column=2, value=f'=IF($A{r}="","",\'Employee Master\'!B{mrow})').font = LINK_FONT
    ws4.cell(row=r, column=3, value=f'=IF($A{r}="","",\'Employee Master\'!C{mrow})').font = LINK_FONT
    ws4.cell(row=r, column=4, value=f'=IF($A{r}="","",IFERROR(VLOOKUP($B$3,\'Monthly Working Days\'!$A$5:$B$24,2,0),""))').font = FORMULA_FONT
    ws4.cell(row=r, column=5, value=f'=IF($A{r}="","",COUNTIFS(\'Attendance Log\'!$B$5:$B$404,$A{r},\'Attendance Log\'!$C$5:$C$404,$B$3))').font = FORMULA_FONT
    ws4.cell(row=r, column=6, value=f'=IF($A{r}="","",D{r}-E{r})').font = FORMULA_FONT
    ws4.cell(row=r, column=7, value=f'=IF($A{r}="","",C{r}*F{r})').font = FORMULA_FONT
    # H: FY leaves used before this month (attendance log count in [FYStart, PrevMonthEnd] + opening leaves)
    ws4.cell(row=r, column=8,
        value=(f'=IF($A{r}="","",COUNTIFS(\'Attendance Log\'!$A$5:$A$404,">="&$P$5,'
               f'\'Attendance Log\'!$A$5:$A$404,"<="&$P$6,\'Attendance Log\'!$B$5:$B$404,$A{r})'
               f'+IFERROR(SUMIFS(\'FY Opening Leaves\'!$C:$C,\'FY Opening Leaves\'!$A:$A,$P$7,'
               f'\'FY Opening Leaves\'!$B:$B,$A{r}),0))')
    ).font = FORMULA_FONT
    # I: FY leaves used including this month = before + this month's absences
    ws4.cell(row=r, column=9, value=f'=IF($A{r}="","",H{r}+E{r})').font = FORMULA_FONT
    # J: excess leaves this month = leaves over 30 after minus leaves over 30 before
    ws4.cell(row=r, column=10, value=f'=IF($A{r}="","",MAX(0,I{r}-30)-MAX(0,H{r}-30))').font = FORMULA_FONT
    # K: basic deduction = excess leaves x (basic / working days this month)
    ws4.cell(row=r, column=11, value=f'=IF($A{r}="","",IFERROR(J{r}*(B{r}/D{r}),0))').font = FORMULA_FONT
    # L: adjusted basic
    ws4.cell(row=r, column=12, value=f'=IF($A{r}="","",B{r}-K{r})').font = FORMULA_FONT
    # M: total payout
    ws4.cell(row=r, column=13, value=f'=IF($A{r}="","",L{r}+G{r})').font = FORMULA_FONT
    for c in [2, 3, 7, 11, 12, 13]:
        ws4.cell(row=r, column=c).number_format = '#,##0;(#,##0);"-"'
    for c in range(1, 14):
        ws4.cell(row=r, column=c).border = BORDER

totrow = 6 + N_EMP
ws4.cell(row=totrow, column=1, value="TOTAL").font = Font(name=FONT_NAME, bold=True)
ws4.cell(row=totrow, column=11, value=f"=SUM(K6:K{totrow-1})").font = Font(name=FONT_NAME, bold=True)
ws4.cell(row=totrow, column=13, value=f"=SUM(M6:M{totrow-1})").font = Font(name=FONT_NAME, bold=True)
ws4.cell(row=totrow, column=11).number_format = '#,##0;(#,##0);"-"'
ws4.cell(row=totrow, column=13).number_format = '#,##0;(#,##0);"-"'
for c in range(1, 14):
    ws4.cell(row=totrow, column=c).border = BORDER

for i, h in enumerate(headers4, start=1):
    ws4.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

set_widths(ws4, [20, 11, 13, 11, 10, 10, 13, 12, 12, 11, 13, 13, 13])
ws4.freeze_panes = "A6"
ws4.row_dimensions[5].height = 45

wb.save("Payroll_Template.xlsx")
print("saved")
